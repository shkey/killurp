#!/usr/bin/env python3

import random
import sys

import requests
from bs4 import BeautifulSoup
from bullet import Bullet, Input, Password, SlidePrompt, YesNo
from openpyxl import Workbook

__VERSION__ = '0.1.1'


class URP:
    def __init__(self):
        self.port = ''
        self.protocol = 'http://'
        self.urp_index = 'newjw.cduestc.cn'
        self.urp_login = '/loginAction.do'
        self.urp_main = '/menu/s_main.jsp'
        self.urp_grade_qb = '/gradeLnAllAction.do?type=ln&oper=qbinfo'
        self.urp_grade_sx = '/gradeLnAllAction.do?type=ln&oper=sxinfo'
        self.urp_grade_fa = '/gradeLnAllAction.do?type=ln&oper=fainfo'
        self.urp_grade_bjg = '/gradeLnAllAction.do?type=ln&oper=bjg'
        self.urp_jxpg_list = '/jxpgXsAction.do?oper=listWj'
        self.urp_jxpg = '/jxpgXsAction.do'
        self.urp_jxpg_page = '/jxpgXsAction.do?oper=wjpg'
        self.comment_words = ['完全 ok', '不错', '可以', '很好', '还行']

    @property
    def index_url(self):
        if self.port:
            return self.protocol + self.urp_index + ':' + self.port
        else:
            return self.protocol + self.urp_index

    @property
    def login_url(self):
        return self.index_url + self.urp_login

    @property
    def main_url(self):
        return self.index_url + self.urp_main

    @property
    def grade_qb_url(self):
        return self.index_url + self.urp_grade_qb

    @property
    def grade_sx_url(self):
        return self.index_url + self.urp_grade_sx

    @property
    def grade_fa_url(self):
        return self.index_url + self.urp_grade_fa

    @property
    def grade_bjg_url(self):
        return self.index_url + self.urp_grade_bjg

    @property
    def jxpg_list_url(self):
        return self.index_url + self.urp_jxpg_list

    @property
    def jxpg_url(self):
        return self.index_url + self.urp_jxpg

    @property
    def jxpg_page_url(self):
        return self.index_url + self.urp_jxpg_page


class Student:
    def __init__(self, account, password):
        self.account = account
        self.password = password
        self._urp = None
        self._session = None

    @property
    def urp(self):
        if self._urp is None:
            self._urp = URP()
        return self._urp

    @property
    def session(self):
        if self._session is None:
            self.make_urp_session()
        return self._session

    def make_urp_session(self):
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3944.0 Safari/537.36',
            'Referer': self.urp.index_url
        }
        data = {
            'zjh': self.account,
            'mm': self.password
        }
        self._session = requests.Session()
        try:
            response = self._session.post(self.urp.login_url, data=data, headers=headers)
            if '学分制综合教务' not in response.text:
                print('账号或密码有误，或者加上端口号再试试？')
                sys.exit()
        except Exception as e:
            print('Exception:', e)
            print('遇到了点小问题，请重试！')
            sys.exit()

    def export_grade(self):
        try:
            print('成绩报表导出中...')
            qb_score_html = self.session.get(self.urp.grade_qb_url)
            qb_score = parse_qbinfo(qb_score_html)
            sx_score_html = self.session.get(self.urp.grade_sx_url)
            sx_score = parse_sxinfo(sx_score_html)
            fa_score_html = self.session.get(self.urp.grade_fa_url)
            fa_score = parse_fainfo(fa_score_html)
            bjg_score_html = self.session.get(self.urp.grade_bjg_url)
            bjg_score = parse_bjg(bjg_score_html)
            save_to_xlsx(qb_score, sx_score, fa_score, bjg_score)
        except Exception as e:
            print('Exception:', e)
            print('遇到了点小问题，请重试！')
            sys.exit()

    def judge_all(self):
        try:
            print('评教中，请稍等...')
            jxpg_jsp = self.session.get(self.urp.jxpg_list_url)
            soup = BeautifulSoup(jxpg_jsp.text, 'lxml')
            course_info_list = soup.select('tr.odd > td > img')
            if len(course_info_list) == 0:
                print('未检测到需要评教的科目...')
            for c in course_info_list:
                c_info = c['name'].split('#@')
                querystring = {
                    'wjbm': c_info[0],
                    'bpr': c_info[1],
                    'pgnr': c_info[5],
                    'oper': 'wjShow',
                    'wjmc': c_info[3].encode('gb2312'),
                    'bprm': c_info[2].encode('gb2312'),
                    'pgnrm': c_info[4].encode('gb2312'),
                    'pageSize': '20',
                    'page': '1',
                    'currentPage': '1',
                    'pageNo': ''
                }
                self.session.post(self.urp.jxpg_page_url, data=querystring)
                querystring = {
                    'wjbm': c_info[0],
                    'bpr': c_info[1],
                    'pgnr': c_info[5],
                    '0000000054': '10_1',
                    '0000000055': '10_1',
                    '0000000056': '5_1',
                    '0000000057': '5_1',
                    '0000000058': '5_1',
                    '0000000059': '5_1',
                    '0000000060': '5_1',
                    '0000000061': '5_1',
                    '0000000062': '5_1',
                    '0000000063': '5_1',
                    '0000000064': '6_1',
                    '0000000065': '7_1',
                    '0000000066': '6_1',
                    '0000000067': '6_1',
                    '0000000068': '5_1',
                    '0000000069': '5_1',
                    '0000000070': '5_1',
                    'zgpj': self.urp.comment_words[random.randint(0, len(self.urp.comment_words) - 1)].encode('gb2312')
                }
                jxpg_submit = self.session.post(self.urp.jxpg_url, data=querystring)
                if jxpg_submit.status_code == requests.codes.get('ok'):
                    print(c_info[4], '评教完成')
            print('程序自动评教完成，请注意自行检查是否真的已评教完成哦～')
        except Exception as e:
            print('Exception:', e)
            print('遇到了点小问题，请重试！')


def parse_qbinfo(score_html):
    soup = BeautifulSoup(score_html.text, 'lxml')
    score_tables = soup.select('#user > tr > td')
    score_list = [['课程号', '课程名', '学分', '课程属性', '成绩']]
    for i in range(0, len(score_tables), 7):
        course_id = score_tables[i].get_text().strip()
        course_name = score_tables[i + 2].get_text().strip()
        course_num = score_tables[i + 4].get_text().strip()
        course_type = score_tables[i + 5].get_text().strip()
        course_score = score_tables[i + 6].get_text().strip()
        score_list.append([course_id, course_name, course_num, course_type, course_score])
    return score_list


def parse_sxinfo(score_html):
    soup = BeautifulSoup(score_html.text, 'lxml')
    score_tables = soup.select('#user > tr > td')
    score_list = [['课程号', '课程名', '学分', '课程属性', '成绩']]
    for i in range(0, len(score_tables), 8):
        course_id = score_tables[i].get_text().strip()
        course_name = score_tables[i + 2].get_text().strip()
        course_num = score_tables[i + 4].get_text().strip()
        course_type = score_tables[i + 5].get_text().strip()
        course_score = score_tables[i + 6].get_text().strip()
        score_list.append([course_id, course_name, course_num, course_type, course_score])
    return score_list


def parse_fainfo(score_html):
    soup = BeautifulSoup(score_html.text, 'lxml')
    score_tables = soup.select('#user > tr > td')
    score_list = [['课程号', '课程名', '学分', '课程属性', '成绩']]
    for i in range(0, len(score_tables), 8):
        course_id = score_tables[i].get_text().strip()
        course_name = score_tables[i + 2].get_text().strip()
        course_num = score_tables[i + 4].get_text().strip()
        course_type = score_tables[i + 5].get_text().strip()
        course_score = score_tables[i + 6].get_text().strip()
        score_list.append([course_id, course_name, course_num, course_type, course_score])
    return score_list


def parse_bjg(score_html):
    soup = BeautifulSoup(score_html.text, 'lxml')
    score_tables = soup.select('#user > tr > td')
    score_list = [['课程号', '课程名', '学分', '课程属性', '成绩', '考试时间']]
    for i in range(0, len(score_tables), 9):
        course_id = score_tables[i].get_text().strip()
        course_name = score_tables[i + 2].get_text().strip()
        course_num = score_tables[i + 4].get_text().strip()
        course_type = score_tables[i + 5].get_text().strip()
        course_score = score_tables[i + 6].get_text().strip()
        course_time = score_tables[i + 7].get_text().strip()
        score_list.append([course_id, course_name, course_num, course_type, course_score, course_time])
    return score_list


def save_to_xlsx(qb_score, sx_score, fa_score, bjg_score):
    try:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = '全部成绩'
        ws2 = wb.create_sheet('课程属性成绩')
        ws3 = wb.create_sheet('方案成绩')
        ws4 = wb.create_sheet('不及格成绩')
        for row in qb_score:
            ws1.append(row)
        for row in sx_score:
            ws2.append(row)
        for row in fa_score:
            ws3.append(row)
        for row in bjg_score:
            ws4.append(row)
        wb.save('scores.xlsx')
        print('成绩报表导出完成，请注意查看...')
    except Exception as e:
        print('Exception:', e)
        print('对不起，好像出了点小问题，要不再试一次？')


def main():
    account_input = Input('请输入你的学号：')
    password_input = Password('请输入你的密码：')
    slide = SlidePrompt([account_input, password_input])
    res = slide.launch()
    account = res[0][1]
    password = res[1][1]
    student = Student(account, password)
    port_confirm_input = YesNo('是否自定义教务系统端口？')
    slide = SlidePrompt([port_confirm_input])
    res = slide.launch()
    port_confirm = res[0][1]
    if port_confirm:
        port_input = Input('请输入你要自定义的端口：')
        slide = SlidePrompt([port_input])
        res = slide.launch()
        student.urp.port = res[0][1]
    while True:
        choices_list = ['成绩导出', '一键评教', '退出']
        choice_input = Bullet('请选择你要进行操作：', choices=choices_list, margin=1)
        slide = SlidePrompt([choice_input])
        res = slide.launch()
        choice = res[0][1]
        if choice == choices_list[0]:
            student.export_grade()
        elif choice == choices_list[1]:
            student.judge_all()
        elif choice == choices_list[2]:
            exit_confirm_input = YesNo('确认退出？')
            slide = SlidePrompt([exit_confirm_input])
            res = slide.launch()
            if res[0][1]:
                sys.exit()


if __name__ == '__main__':
    main()
