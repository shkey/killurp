#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import prettytable
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

import student


def get_qbinfo(score_html):
    # 将得到的教务系统的源码转换成BeautifulSoup对象，并指定解析器为lxml
    soup = BeautifulSoup(score_html.text, "lxml")
    # 通过css选择器选出要用的数据
    score_tables = soup.select("#user > tr > td")
    score_list = []
    score_list.append(["课程号", "课程名", "学分", "课程属性", "成绩"])
    for i in range(0, len(score_tables), 7):
        course_id = score_tables[i].get_text().strip()
        course_name = score_tables[i + 2].get_text().strip()
        course_num = score_tables[i + 4].get_text().strip()
        course_type = score_tables[i + 5].get_text().strip()
        course_score = score_tables[i + 6].get_text().strip()
        score_list.append([course_id, course_name, course_num, course_type, course_score])
    return score_list

def get_sxinfo(score_html):
    # 将得到的教务系统的源码转换成BeautifulSoup对象，并指定解析器为lxml
    soup = BeautifulSoup(score_html.text, "lxml")
    # 通过css选择器选出要用的数据
    score_tables = soup.select("#user > tr > td")
    score_list = []
    score_list.append(["课程号", "课程名", "学分", "课程属性", "成绩"])
    for i in range(0, len(score_tables), 8):
        course_id = score_tables[i].get_text().strip()
        course_name = score_tables[i + 2].get_text().strip()
        course_num = score_tables[i + 4].get_text().strip()
        course_type = score_tables[i + 5].get_text().strip()
        course_score = score_tables[i + 6].get_text().strip()
        score_list.append([course_id, course_name, course_num, course_type, course_score])
    return score_list

def get_fainfo(score_html):
    # 将得到的教务系统的源码转换成BeautifulSoup对象，并指定解析器为lxml
    soup = BeautifulSoup(score_html.text, "lxml")
    # 通过css选择器选出要用的数据
    score_tables = soup.select("#user > tr > td")
    score_list = []
    score_list.append(["课程号", "课程名", "学分", "课程属性", "成绩"])
    for i in range(0, len(score_tables), 8):
        course_id = score_tables[i].get_text().strip()
        course_name = score_tables[i + 2].get_text().strip()
        course_num = score_tables[i + 4].get_text().strip()
        course_type = score_tables[i + 5].get_text().strip()
        course_score = score_tables[i + 6].get_text().strip()
        score_list.append([course_id, course_name, course_num, course_type, course_score])
    return score_list

def get_bjg(score_html):
    # 将得到的教务系统的源码转换成BeautifulSoup对象，并指定解析器为lxml
    soup = BeautifulSoup(score_html.text, "lxml")
    # 通过css选择器选出要用的数据
    score_tables = soup.select("#user > tr > td")
    score_list = []
    score_list.append(["课程号", "课程名", "学分", "课程属性", "成绩", "考试时间"])
    for i in range(0, len(score_tables), 9):
        course_id = score_tables[i].get_text().strip()
        course_name = score_tables[i + 2].get_text().strip()
        course_num = score_tables[i + 4].get_text().strip()
        course_type = score_tables[i + 5].get_text().strip()
        course_score = score_tables[i + 6].get_text().strip()
        course_time = score_tables[i + 7].get_text().strip()
        score_list.append([course_id, course_name, course_num, course_type, course_score, course_time])
    return score_list

def print_qbinfo(choice_html):
    # 将得到的教务系统的源码转换成BeautifulSoup对象，并指定解析器为lxml
    soup = BeautifulSoup(choice_html.text, "lxml")
    # 通过css选择器选出要用的数据
    score_tables = soup.select("#user > tr > td")
    pt_score = prettytable.PrettyTable(["课程号", "课程名", "学分", "课程属性", "成绩"])
    # 设置prettytable靠左对齐
    pt_score.align = "l"
    for i in range(0, len(score_tables), 7):
        course_id = score_tables[i].get_text().strip()
        course_name = score_tables[i + 2].get_text().strip()
        course_num = score_tables[i + 4].get_text().strip()
        course_type = score_tables[i + 5].get_text().strip()
        course_score = score_tables[i + 6].get_text().strip()
        pt_score.add_row([course_id, course_name, course_num, course_type, course_score])
    print("全部成绩")
    print(pt_score)

def print_sxinfo(choice_html):
    # 将得到的教务系统的源码转换成BeautifulSoup对象，并指定解析器为lxml
    soup = BeautifulSoup(choice_html.text, "lxml")
    # 通过css选择器选出要用的数据
    score_tables = soup.select("#user > tr > td")
    pt_score = prettytable.PrettyTable(["课程号", "课程名", "学分", "课程属性", "成绩"])
    # 设置prettytable靠左对齐
    pt_score.align = "l"
    for i in range(0, len(score_tables), 8):
        course_id = score_tables[i].get_text().strip()
        course_name = score_tables[i + 2].get_text().strip()
        course_num = score_tables[i + 4].get_text().strip()
        course_type = score_tables[i + 5].get_text().strip()
        course_score = score_tables[i + 6].get_text().strip()
        pt_score.add_row([course_id, course_name, course_num, course_type, course_score])
    print("课程属性成绩")
    print(pt_score)

def print_fainfo(choice_html):
    # 将得到的教务系统的源码转换成BeautifulSoup对象，并指定解析器为lxml
    soup = BeautifulSoup(choice_html.text, "lxml")
    # 通过css选择器选出要用的数据
    score_tables = soup.select("#user > tr > td")
    pt_score = prettytable.PrettyTable(["课程号", "课程名", "学分", "课程属性", "成绩"])
    # 设置prettytable靠左对齐
    pt_score.align = "l"
    for i in range(0, len(score_tables), 8):
        course_id = score_tables[i].get_text().strip()
        course_name = score_tables[i + 2].get_text().strip()
        course_num = score_tables[i + 4].get_text().strip()
        course_type = score_tables[i + 5].get_text().strip()
        course_score = score_tables[i + 6].get_text().strip()
        pt_score.add_row([course_id, course_name, course_num, course_type, course_score])
    print("方案成绩")
    print(pt_score)

def print_bjg(choice_html):
    # 将得到的教务系统的源码转换成BeautifulSoup对象，并指定解析器为lxml
    soup = BeautifulSoup(choice_html.text, "lxml")
    # 通过css选择器选出要用的数据
    score_tables = soup.select("#user > tr > td")
    pt_score = prettytable.PrettyTable(["课程号", "课程名", "学分", "课程属性", "成绩", "考试时间"])
    # 设置prettytable靠左对齐
    pt_score.align = "l"
    for i in range(0, len(score_tables), 9):
        course_id = score_tables[i].get_text().strip()
        course_name = score_tables[i + 2].get_text().strip()
        course_num = score_tables[i + 4].get_text().strip()
        course_type = score_tables[i + 5].get_text().strip()
        course_score = score_tables[i + 6].get_text().strip()
        course_time = score_tables[i + 7].get_text().strip()
        pt_score.add_row([course_id, course_name, course_num, course_type, course_score, course_time])
    print("不及格成绩")
    print(pt_score)

def save_to_xlsx(qb_score, sx_score, fa_score, bjg_score):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "全部成绩"
    ws2 = wb.create_sheet("课程属性成绩")
    ws3 = wb.create_sheet("方案成绩")
    ws4 = wb.create_sheet("不及格成绩")
    for row in qb_score:
        ws1.append(row)
    for row in sx_score:
        ws2.append(row)
    for row in fa_score:
        ws3.append(row)
    for row in bjg_score:
        ws4.append(row)
    wb.save('scores.xlsx')
    try:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "全部成绩"
        ws2 = wb.create_sheet("课程属性成绩")
        ws3 = wb.create_sheet("方案成绩")
        ws4 = wb.create_sheet("不及格成绩")
        for row in qb_score:
            ws1.append(row)
        for row in sx_score:
            ws2.append(row)
        for row in fa_score:
            ws3.append(row)
        for row in bjg_score:
            ws4.append(row)
        wb.save('scores.xlsx')
        print("成绩报表导出完成，请注意查看")
    except:
        print("对不起，好像出了点小问题，要不再试一次？")

def main(stu):
    # 构造一个列表，用来存放查询成绩的类型
    score_urls = [
        "http://newjw.cduestc.cn/gradeLnAllAction.do?type=ln&oper=qbinfo", "http://newjw.cduestc.cn/gradeLnAllAction.do?type=ln&oper=sxinfo", "http://newjw.cduestc.cn/gradeLnAllAction.do?type=ln&oper=fainfo", "http://newjw.cduestc.cn/gradeLnAllAction.do?type=ln&oper=bjg"
    ]
    pt = prettytable.PrettyTable(["查询类型"])
    # 设置prettytable靠左对齐
    pt.align = "l"
    pt.add_row(["0、导出成绩报表"])
    pt.add_row(["1、全部成绩查询"])
    pt.add_row(["2、课程属性成绩查询"])
    pt.add_row(["3、方案成绩查询"])
    pt.add_row(["4、不及格成绩查询"])
    print(pt)
    try:
        choice = input("请按序号输入你要进行的操作\r\n# ").strip()
        if choice.isdigit():
            choice = int(choice)
            score_html = stu.sess.get(score_urls[choice - 1], timeout=10)
            if choice == 1:
                print_qbinfo(score_html)
            elif choice == 2:
                print_sxinfo(score_html)
            elif choice == 3:
                print_fainfo(score_html)
            elif choice == 4:
                print_bjg(score_html)
            elif choice == 0:
                qb_score_html = stu.sess.get(score_urls[0], timeout=10)
                qb_score = get_qbinfo(qb_score_html)
                sx_score_html = stu.sess.get(score_urls[1], timeout=10)
                sx_score = get_sxinfo(sx_score_html)
                fa_score_html = stu.sess.get(score_urls[2], timeout=10)
                fa_score = get_fainfo(fa_score_html)
                bjg_score_html = stu.sess.get(score_urls[3], timeout=10)
                bjg_score = get_bjg(bjg_score_html)
                save_to_xlsx(qb_score, sx_score, fa_score, bjg_score)
            print("成绩查询完成，已为你回退到主菜单")
        else:
            print("你的输入有误，已为你回退到主菜单")
            return 0
    except:
        print("遇到了点小问题，请重试")


if __name__ == "__main__":
    stu = student.Student("123", "test")
    stu.login()
    main(stu)
